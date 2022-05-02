
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, Color
import os 

class Analyse:
    def __init__(self,file_name,folder_directory):

        self.file_name = file_name
        self.folder_directory = folder_directory



    #Provides the number of files in the folder 
    def number_of_files(self):
        counter = 0 #keep a count of all files found
        location= self.folder_directory # will get folder location here where many xlsx files are present
        for file in os.listdir(location):
            try:
                if file.endswith(".xlsx"):
                    counter = counter+1
            except Exception as e:
                raise e
                print ("No files found here!")
        return counter


    # Creating list of required values from all workbooks
    def collect_data(self):
        os.chdir(self.folder_directory) #moving to directory where all workbooks are available
        self.count = self.number_of_files()
        print(self.count)
        self.data = []
        self.measurement = []
        self.depth_list = []
        self.roughness_list = []
        for workbook in range(1,self.count):   #when file count starts from 1 iterate till count
            self.data.append(workbook)

            if (workbook<10):
                wb = load_workbook(str(self.file_name)+'_000'+str(workbook)+'.xlsx')
            if (workbook<100 and workbook>=10 ):
                wb = load_workbook(str(self.file_name)+'_00'+str(workbook)+'.xlsx')
            if (workbook >= 100): 
                wb = load_workbook(str(self.file_name)+'_0'+str(workbook)+'.xlsx')

            #To obtain the sheet number in the workbook
            self.sheet_number = len(wb.worksheets)
            if (self.sheet_number == 4):
                #To obtain all depth values
                self.ws = wb['Profile']
                #to find the average of the depth
                self.depth_list.extend((self.ws['E85'].value,self.ws['E86'].value,self.ws['E87'].value,self.ws['E88'].value)) # make a list using all depth values from different measurement lines
                self.depth_list.sort(reverse = True) # sort the list in descending order

                #iteration removes a number if the difference between the depth value is greater than 10
                for i in range(len(self.depth_list)-1):
                    self.difference = self.depth_list[0]-self.depth_list[i]
                    if self.difference>=15: self.depth_list.remove(self.depth_list[i])
                self.depth = self.average(self.depth_list)
                self.data.append(round(self.depth,3))

                #To obtain areal roughness roughness values
                self.ws = wb["Areal roughness"]
                self.roughness = ((self.ws['C24'].value))
                self.data.append(round(self.roughness,3))

            if(self.sheet_number == 3):
                #To obtain all depth values
                self.ws = wb['Profile']
                #to find the average of the depth
                self.depth_list.extend((self.ws['E85'].value,self.ws['E86'].value,self.ws['E87'].value,self.ws['E88'].value)) # make a list using all depth values from different measurement lines
                self.depth_list.sort(reverse = True) # sort the list in descending order

                #iteration removes a number if the difference between the depth value is greater than 10
                for i in range(len(self.depth_list)-1):
                    self.difference = self.depth_list[0]-self.depth_list[i]
                    if self.difference>=15: self.depth_list.remove(self.depth_list[i])
                self.depth = self.average(self.depth_list) 
                self.data.append(round(self.depth,3))


            self.measurement.append(self.data)
            self.data = []
            self.depth_list = []
            self.roughness_list = []
        return self.measurement


    # to get average of a list
    def average(self,lst):
        return sum(lst) / len(lst)

    # Creating a new excel sheet to store required values
    def create_master_sheet(self):
        self.wb_new = Workbook()
        self.ws_new = self.wb_new.active
        self.ws_new.title = "Master sheet"
        self.headings = ['No.','depth(μm)','roughness(Sq[µm])']
        self.ws_new.append(self.headings)


        #creating a style for the headings
        self.highlight = NamedStyle(name="highlight")
        self.highlight.font = Font(bold=True, size=10)

        #applyign style to the headings
        self.ws_new['A1'].style = self.highlight
        self.ws_new['B1'].style = 'highlight'
        self.ws_new['C1'].style = 'highlight'

        self.data = self.collect_data()


        for i in self.data:
            print(i)  
            self.ws_new.append(i) 

        self.wb_new.save('Measurements.xlsx')

        
