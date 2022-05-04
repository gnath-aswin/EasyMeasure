
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, Color
import os 
import pandas as pd

class Analyse:
    def __init__(self,folder_directory):

        self.file_name = os.path.basename(folder_directory.rsplit('_',1)[0])
        self.folder_directory = os.path.dirname(os.path.abspath(folder_directory)) #only works in windows
        self.selected_parameters = []
        self.depth_cell_number = ''
        self.areal_roughness_cell_number = ''
        self.surface_roughness_cell_number = ''


    #Provides the number of files in the folder 
    def number_of_files(self):
        counter = 0 #keep a count of all files found
        location= self.folder_directory # will get folder location here where many xlsx files are present
        for file in os.listdir(location):
            try:
                if file.endswith(".xlsx"):
                    counter = counter+1
            except Exception as e:
                raise e
                print ("No files found here!")
        return counter


    # Creating list of required values from all workbooks

    def collect_data(self):
        os.chdir(self.folder_directory) #moving to directory where all workbooks are available
        self.count = self.number_of_files()
        print(self.count)
        self.depth_ = []
        self.depth_list = []
        self.roughness_ = []
        self.roughness_list = []

        #information from user --> required parameters and cell number



        for workbook in range(1,self.count):   #when file count starts from 1 iterate till count
            #self.data.append(workbook)

            if (workbook<10):
                wb = load_workbook(str(self.file_name)+'_000'+str(workbook)+'.xlsx')
            if (workbook<100 and workbook>=10 ):
                wb = load_workbook(str(self.file_name)+'_00'+str(workbook)+'.xlsx')
            if (workbook >= 100): 
                wb = load_workbook(str(self.file_name)+'_0'+str(workbook)+'.xlsx')


            for parameter in self.selected_parameters:
                if (parameter == 'depth'):
                    #open profile sheet
                    try:self.ws = wb['Profile']
                    except: 
                        self.roughness_list.append(0)
                        break

                    #need to split parameter and corresponding cell numbers
                    self.depth_cell_numbers = self.depth_cell_number.split(",")

                    #create a list a depth values from user given excel cell number
                    for i in range(int(len(self.depth_cell_numbers))):
                        self.depth_.append(self.ws[self.depth_cell_numbers[i]].value)

                    # sort the list in descending order   
                    self.depth_.sort(reverse = True) 

                    #iteration removes a number if the difference between the depth value is greater than 15
                    for i in range(len(self.depth_)-1):
                        self.difference = self.depth_[0]-self.depth_[i]
                        if self.difference>=15: self.depth_.remove(self.depth_[i])
                    self.depth = self.average(self.depth_)
                    self.depth_list.append(round(self.depth,3))

                if(parameter == 'areal roughness'): 
                    #To obtain areal roughness roughness values
                    try:self.ws = wb['Areal roughness']
                    except: 
                        self.roughness_list.append(0)
                        break

                    #need to split parameter and corresponding cell numbers
                    self.areal_roughness_cell_numbers = self.areal_roughness_cell_number.split(",")

                    #create a list a depth values from user given excel cell number
                    for i in range(int(len(self.areal_roughness_cell_numbers))):
                        self.roughness_.append(self.ws[self.areal_roughness_cell_numbers[i]].value)

                    # sort the list in descending order   
                    self.roughness_.sort(reverse = True) 

                    #iteration removes a number if the difference between the depth value is greater than 15
                    for i in range(len(self.roughness_)-1):
                        self.difference = self.roughness_[0]-self.roughness_[i]
                        if self.difference>=15: self.roughness_.remove(self.roughness_[i])
                    self.roughness = self.average(self.roughness_)
                    self.roughness_list.append(round(self.roughness,3))

      
            self.depth_ = []
            self.roughness_ = []

        if (self.selected_parameters == ['depth','areal roughness']):
            self.df = pd.DataFrame({'depth(μm)':self.depth_list, 'roughness(Sq[µm])':self.roughness_list})

        if (self.selected_parameters == ['depth']):
            self.df = pd.DataFrame({'depth(μm)':self.depth_list})

        return self.df

    # to get average of a list
    def average(self,lst):
        return sum(lst) / len(lst)

    # Creating a new excel sheet to store required values
    def create_master_sheet(self,row,column):

        data  = self.collect_data()

        # if (self.selected_parameters == ['depth','areal roughness']):
        #     data.index += 1            #to start index in pandas dataframe from 1
        #     if row != 0 and column != 0:
        #         depth_2d = pd.DataFrame(data[['depth(μm)']].to_numpy().reshape(row,column)) #reshape(row,column)
        #         depth_2d.index += 1
        #         roughness_2d = pd.DataFrame(data[['roughness(Sq[µm])']].to_numpy().reshape(row,column)) #reshape(row,column)
        #         roughness_2d.index += 1 
        #     with pd.ExcelWriter('output.xlsx') as writer:  
        #         data.to_excel(writer, sheet_name='Main')
        #         if row != 0 and column != 0:
        #             depth_2d.to_excel(writer, sheet_name='Depth')
        #             roughness_2d.to_excel(writer, sheet_name='Roughness')
        
        for parameter in self.selected_parameters:
            data.index += 1            #to start index in pandas dataframe from 1
            if row != 0 and column != 0:
                parameter_2d = pd.DataFrame(data[[parameter]].to_numpy().reshape(row,column)) #reshape(row,column)
                parameter_2d.index += 1
            with pd.ExcelWriter('output.xlsx') as writer:  
                data.to_excel(writer, sheet_name='Main')
                if row != 0 and column != 0:
                    parameter_2d.to_excel(writer, sheet_name=parameter)

        
